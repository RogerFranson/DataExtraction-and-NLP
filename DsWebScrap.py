import openpyxl
import requests
import os
import nltk.tokenize
from nltk.tokenize import word_tokenize
from bs4 import BeautifulSoup
import re
import xlsxwriter


# Give the location of the file
path = "Input.xlsx"

workbook = xlsxwriter.Workbook("OutputDS.xlsx")
worksheet = workbook.add_worksheet()
title = [
    "URL_ID",
    "URL",
    "POSITIVE SCORE",
    "NEGATIVE SCORE",
    "POLARITY SCORE",
    "SUBJECTIVITY SCORE",
    "AVG SENTENCE LENGTH",
    "PERCENTAGE OF COMPLEX WORDS",
    "FOG INDEX",
    "AVG NUMBER OF WORDS PER SENTENCE",
    "COMPLEX WORD COUNT",
    "WORD COUNT",
    "SYLLABLE PER WORD",
    "PERSONAL PRONOUNS",
    "AVG WORD LENGTH",
]
for i in range(0, len(title)):
    worksheet.write(0, i, title[i])

# returns article data a tokens
def url_par(article_path, r_val):
    def read_Article(a_path):
        try:

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
            }
            response = requests.get(
                url=a_path,
                headers=headers,
            )

            soup = BeautifulSoup(response.text, "html.parser")
            # xpath //div[@class="td-post-content"]/p
            article_Arr = ""

            for points in soup.find("div", class_="td-post-content").find_all("p"):
                point = str(points.text)
                article_Arr += point
                No_sentence = article_Arr.count(".")  # len(article_Arr.split("."))
            article_Arr = re.sub(r"[^\w\s]", "", article_Arr)

            return article_Arr, No_sentence

        except Exception as e:

            url_id = 35 + r_val
            url = article_path
            worksheet.write(r_val - 1, 0, url_id)
            worksheet.write(r_val - 1, 1, url)
            worksheet.write(r_val - 1, 2, "No Article Data found")
            print(e)

    # returns the Processing token values from the file diretory
    def get_fileData(Directory_name):
        dir_path = r"{0}".format(Directory_name)
        no_file = len(os.listdir(dir_path))
        f_name = []
        for i in range(0, no_file):
            f_name.append((os.listdir(dir_path)[i].split("."))[0])
            y = f_name[i]
            data_arr.update({y: ""})
            path = os.listdir(dir_path)[i]
            f = open(dir_path + "\{0}".format(path), "r")
            f_name[i] = f.read()
            data_arr.update({y: word_tokenize(f_name[i])})
            # print(f_name[i])
        return data_arr

    def syllables(word):
        count_s = 0
        vowels = "aeiouy"
        word = word.lower()
        if word[0] in vowels:
            count_s += 1
        for index in range(1, len(word)):
            if word[index] in vowels and word[index - 1] not in vowels:
                count_s += 1
        if word.endswith("e") or word.endswith("ed") or word.endswith("es"):
            count_s -= 1
        if word.endswith("le"):
            count_s += 1
        if count_s == 0:
            count_s += 1
        return count_s

    x = read_Article(article_path)

    def get_pronoun(x):
        pronounRegex = re.compile(r"\b(I|we|my|ours|(?-i:us))\b", re.I)
        pronouns = pronounRegex.findall(x)
        return len(pronouns)

    pronoun_count = get_pronoun(x[0])
    # pronoun_count = 0
    urlDat = word_tokenize(x[0])
    art_len = len(urlDat)
    print(article_path)
    No_sentence = x[1]
    data_arr = {}

    get_fileData("MasterDictionary")
    get_fileData("StopWords")
    pos_num = 0
    neg_num = 0
    for x in data_arr.keys():
        if "StopWords" in x:
            p_d = data_arr[x]
            for word in p_d:
                temp = urlDat.count(word)
                for i in range(0, temp):
                    urlDat.remove(word)
        elif "negative" in x:
            p_d = data_arr[x]
            for word in p_d:
                if word in urlDat:
                    neg_num += 1
                    # print(word)
        elif "positive" in x:
            p_d = data_arr[x]
            for word in p_d:
                if word in urlDat:
                    pos_num += 1
                    # print(word)

    complexWords_count = 0
    sy_count = 0
    for sy in urlDat:
        sy_count += syllables(sy)
        if syllables(sy) > 2:
            complexWords_count += 1

    url_id = 35 + r_val
    url = article_path
    p_Score = int(pos_num)
    n_ScoreE = neg_num
    polarity = ((pos_num - neg_num) / (pos_num + neg_num)) + 0.000001
    subjective = ((pos_num + neg_num) / len(urlDat)) + 0.000001
    avg_length_wrd = art_len / No_sentence
    Complex_percet = complexWords_count / art_len
    fog_index = 0.4 * (avg_length_wrd + Complex_percet)
    word_p_sentence = art_len / No_sentence
    Comple_count = complexWords_count
    Word_count = len(urlDat)
    Sy_PER_WORD = sy_count
    pl_PRONOUNS = pronoun_count
    avg_wrd_LENGTH = len(x[0]) / art_len
    final_data = [
        url_id,
        url,
        p_Score,
        n_ScoreE,
        polarity,
        subjective,
        avg_length_wrd,
        Complex_percet,
        fog_index,
        word_p_sentence,
        Comple_count,
        Word_count,
        Sy_PER_WORD,
        pl_PRONOUNS,
        avg_wrd_LENGTH,
    ]

    for m in range(0, 15):
        worksheet.write(r_val - 1, m, final_data[m])


wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
for i in range(2, sheet_obj.max_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=2)
    r_val = i
    article_path = cell_obj.value
    try:
        url_par(article_path, r_val)
    except Exception as e:
        url_id = 35 + r_val
        url = article_path
        worksheet.write(r_val - 1, 0, url_id)
        worksheet.write(r_val - 1, 1, url)
        worksheet.write(r_val - 1, 2, "No Article Data found")
        print(e)

workbook.close()
